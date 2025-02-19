import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import argparse

def log_processing_info(log_file, message):
    """
    Scrie în fișierul 'log_file' un mesaj, cu data/ora curentă.
    Creează fișierul dacă nu există.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{now}] {message}\n")

def find_csv_pairs(csv_folder):
    """
    Identifică perechile de fișiere CSV (1M și 3M) pe baza numelui comun.
    Returnează un dict de forma:
    {
      'Nume Comun': {
          '1M': 'cale/catre/..., 1M.csv',
          '3M': 'cale/catre/..., 3M.csv'
       },
      ...
    }
    """
    csv_files = os.listdir(csv_folder)
    csv_pairs = {}
    
    for file in csv_files:
        if file.endswith(".csv"):
            base_name = file.rsplit(", ", 1)[0]  # Eliminăm partea finală (1M sau 3M)
            if "1M" in file:
                csv_pairs.setdefault(base_name, {})["1M"] = os.path.join(csv_folder, file)
            elif "3M" in file:
                csv_pairs.setdefault(base_name, {})["3M"] = os.path.join(csv_folder, file)
    
    return csv_pairs

def clear_last_row_formulas_lunar(ws, last_data_row):
    """
    Elimină formulele din coloanele specifice pentru ultimul rând generat 
    pentru date lunare (1M).
    """
    columns_to_clear = ["J", "K", "L", "M", "N", "O", 
                        "AL", "AM", "AN", "AO", "AP", "AQ"]
    for col in columns_to_clear:
        ws[f"{col}{last_data_row}"].value = None

def clear_last_row_formulas_quarter(ws, last_data_row_quarter):
    """
    Elimină formulele din coloanele "Y", "Z", "AA", "AB", "AC", "AD"
    pentru ultimul rând din datele trimestriale (3M).
    """
    columns_to_clear = ["Y", "Z", "AA", "AB", "AC", "AD"]
    for col in columns_to_clear:
        ws[f"{col}{last_data_row_quarter}"].value = None

def remove_extra_rows(ws, num_rows):
    """
    Elimină rândurile în plus din fișierul Excel dacă sunt mai multe decât 
    în fișierul CSV 1M (num_rows).
    """
    max_existing_rows = ws.max_row
    if max_existing_rows > num_rows + 1:
        ws.delete_rows(num_rows + 2, max_existing_rows - num_rows - 1)

def remove_extra_rows_quarter(ws, num_rows_quarter):
    """
    Elimină valorile în plus din coloanele T..AH dacă fișierul 3M 
    are mai puține rânduri decât template-ul.
    """
    columns_to_clear = [
        "R", "S", "T", "U", "V", "W", "X", "Y", "Z", 
        "AA", "AB", "AC", "AD"
    ]
    max_existing_rows = ws.max_row
    
    # Golim conținutul (nu ștergem rândurile) după num_rows_quarter
    for row in range(num_rows_quarter + 2, max_existing_rows + 1):
        for col in columns_to_clear:
            ws[f"{col}{row}"].value = None

def process_csv_to_xlsx(csv_folder, template_path, output_folder, log_file="process_log.txt"):
    """
    Completează template.xlsx cu datele din fișierele CSV (1M și 3M),
    fără a modifica structura acestuia. 
    Scrie log-uri în 'log_file'.
    """
    csv_pairs = find_csv_pairs(csv_folder)
    
    # Creăm folderul de output dacă nu există
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Log: începem procesarea
    log_processing_info(log_file, f"Start procesare fișiere.'")
    
    for base_name, files in csv_pairs.items():
        # Verificăm dacă avem și fișier 1M, și fișier 3M
        if "1M" not in files or "3M" not in files:
            error_msg = f"Eroare: Fișierele CSV 1M / 3M pentru '{base_name}' nu sunt complete."
            print(error_msg)
            log_processing_info(log_file, error_msg)
            # Continuăm cu următorul 'base_name'
            continue
        
        try:
            # Încarcă șablonul Excel
            wb = load_workbook(template_path)
            ws = wb.active
            
            # Redenumește foaia de lucru
            ws.title = base_name
            
            # Citește datele din CSV-uri
            df_lunar = pd.read_csv(files["1M"])
            df_quarter = pd.read_csv(files["3M"])
            
            num_rows = len(df_lunar)
            num_rows_quarter = len(df_quarter)
            
            # Inserăm datele 1M (exemplu: coloanele A și B)
            for index, row in df_lunar.iterrows():
                ws.cell(row=index+2, column=1, value=row.iloc[0])  # data 1M
                ws.cell(row=index+2, column=2, value=row.iloc[1])  # valoare 1M
            
            # Inserăm datele 3M (exemplu: coloanele R și S)
            for index, row in df_quarter.iterrows():
                if index < num_rows:  
                    ws.cell(row=index+2, column=16, value=row.iloc[0])  # data 3M
                    ws.cell(row=index+2, column=17, value=row.iloc[1])  # valoare 3M
            
            # Eliminăm rândurile și formulele în plus
            remove_extra_rows(ws, num_rows)
            remove_extra_rows_quarter(ws, num_rows_quarter)
            clear_last_row_formulas_lunar(ws, num_rows + 1)
            clear_last_row_formulas_quarter(ws, num_rows_quarter + 1)
            
            # Salvează fișierul XLSX rezultat
            output_file = os.path.join(output_folder, f"{base_name}.xlsx")
            wb.save(output_file)
            
            msg_ok = f"✔ Fișier completat: {base_name}.xlsx"
            print(msg_ok)
            log_processing_info(log_file, msg_ok)
        
        except Exception as e:
            # Dacă apare vreo eroare la procesare, o logăm și continuăm
            error_msg = f"Eroare la procesarea '{base_name}': {e}"
            print(error_msg)
            log_processing_info(log_file, error_msg)
    
    # Log: finalizare
    log_processing_info(log_file, "Final procesare fișiere.")

# ------------------------------------------------
# EXEMPLU de utilizare (script standalone):
# ------------------------------------------------
if __name__ == "__main__":

    # 1) Încarcă argumentele din linia de comandă
    parser = argparse.ArgumentParser(description="Procesare CSV în XLSX.")
    parser.add_argument(
        "--csv-folder",
        type=str,
        default=None,
        help="Folderul unde sunt fișierele CSV."
    )
    parser.add_argument(
        "--template-path",
        type=str,
        default=None,
        help="Calea către fișierul template XLSX."
    )
    parser.add_argument(
        "--output-folder",
        type=str,
        default=None,
        help="Folderul unde se vor salva fișierele XLSX rezultate."
    )
    parser.add_argument(
        "--log-file",
        type=str,
        default=None,
        help="Numele/locația fișierului de log."
    )
    args = parser.parse_args()

    # 2) Preia și din variabile de mediu (dacă nu vin din CLI):
    #    CSV_FOLDER, TEMPLATE_PATH, OUTPUT_FOLDER, LOG_FILE
    csv_folder_env = os.getenv("CSV_FOLDER", "csv")
    template_path_env = os.getenv("TEMPLATE_PATH", "template/template.xlsx")
    output_folder_env = os.getenv("OUTPUT_FOLDER", "output")
    log_file_env = os.getenv("LOG_FILE", "process_log.txt")

    # 3) Alege ordinea priorităților:
    #    CLI > Environment Variables > Valorile default
    csv_folder = args.csv_folder if args.csv_folder else csv_folder_env
    template_path = args.template_path if args.template_path else template_path_env
    output_folder = args.output_folder if args.output_folder else output_folder_env
    log_file = args.log_file if args.log_file else log_file_env

    # --------------------------------------------------------
    # Liniile următoare erau în codul original - le comentăm,
    # dar nu le ștergem (cerința ta).
    # --------------------------------------------------------
    # csv_folder = "csv"
    # template_path = "template/template.xlsx"
    # output_folder = "output"
    # log_file = "process_log.txt"
    #
    # process_csv_to_xlsx(csv_folder, template_path, output_folder, log_file)

    # --------------------------------------------------------
    # Apel final cu parametrii obținuți din CLI / ENV
    # --------------------------------------------------------
    process_csv_to_xlsx(
        csv_folder=csv_folder,
        template_path=template_path,
        output_folder=output_folder,
        log_file=log_file
    )