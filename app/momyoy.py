import os
import re
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def log_processing_info(log_file, message):
    """
    Scrie în fișierul 'log_file' un mesaj, cu data/ora curentă.
    Creează fișierul dacă nu există.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{now}] {message}\n")

def find_csv_pairs(csv_folder):
    """
    Identifică perechile de fișiere CSV (1M și 3M) pe baza numelui comun.
    Returnează un dict de forma:
    {
      'Nume Comun': {
          '1M': 'cale/catre/..., 1M.csv',
          '3M': 'cale/catre/..., 3M.csv'
       },
      ...
    }
    """
    csv_files = os.listdir(csv_folder)
    csv_pairs = {}
    
    for file in csv_files:
        if file.endswith(".csv"):
            # Regex pentru a extrage numele principal și tipul indicatorului (MM sau YY)
            match = re.match(r"(ECONOMICS_[A-Z]+)(MM|YY),\s*\dM\.csv$", file)
            if match:
                base_name, indicator_type = match.groups()  # Extragem numele + tipul indicatorului
                
                # Adăugăm fișierul în dicționar, grupat după numele principal
                csv_pairs.setdefault(base_name, {})[indicator_type] = os.path.join(csv_folder, file)

    # Filtrăm doar acele intrări care au atât MM, cât și YY
    csv_pairs = {k: v for k, v in csv_pairs.items() if "MM" in v and "YY" in v}

    return csv_pairs

def clear_last_row_formulas_lunar(ws, last_data_row):
    """
    Elimină formulele din coloanele specifice pentru ultimul rând generat 
    pentru date lunare (1M).
    """
    columns_to_clear = ["I", "J", "K", "L", "M", "N"]
    for col in columns_to_clear:
        ws[f"{col}{last_data_row}"].value = None

def clear_last_row_formulas_quarter(ws, last_data_row_quarter):
    """
    Elimină formulele din coloanele
    pentru ultimul rând din datele trimestriale (3M).
    """
    columns_to_clear = ["X", "Y", "Z", "AA", "AB", "AC"]
    for col in columns_to_clear:
        ws[f"{col}{last_data_row_quarter}"].value = None

def remove_extra_rows(ws, num_rows):
    """
    Elimină rândurile în plus din fișierul Excel dacă sunt mai multe decât 
    în fișierul CSV 1M (num_rows).
    """
    max_existing_rows = ws.max_row
    if max_existing_rows > num_rows + 1:
        ws.delete_rows(num_rows + 2, max_existing_rows - num_rows - 1)

def remove_extra_rows_quarter(ws, num_rows_quarter):
    """
    Elimină valorile în plus din coloane dacă fișierul 3M 
    are mai puține rânduri decât template-ul.
    """
    columns_to_clear = [
        "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", 
        "AA", "AB", "AC"
    ]
    max_existing_rows = ws.max_row
    
    # Golim conținutul (nu ștergem rândurile) după num_rows_quarter
    for row in range(num_rows_quarter + 2, max_existing_rows + 1):
        for col in columns_to_clear:
            ws[f"{col}{row}"].value = None

def process_csv_to_xlsx(csv_folder, template_path, output_folder, log_file="process_log.txt"):
    """
    Completează template.xlsx cu datele din fișierele CSV (1M și 3M),
    fără a modifica structura acestuia. 
    Scrie log-uri în 'log_file'.
    """
    csv_pairs = find_csv_pairs(csv_folder)
    
    # Creăm folderul de output dacă nu există
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Log: începem procesarea
    log_processing_info(log_file, f"Start procesare fișiere din folderul '{csv_folder}'")
    
    for base_name, files in csv_pairs.items():
        # Verificăm dacă avem și fișier 1M, și fișier 3M
        if "MM" not in files or "YY" not in files:
            error_msg = f"Eroare: Fișierele CSV 1M / 3M pentru '{base_name}' nu sunt complete."
            print(error_msg)
            log_processing_info(log_file, error_msg)
            # Continuăm cu următorul 'base_name'
            continue
        
        try:
            # Încarcă șablonul Excel
            wb = load_workbook(template_path)
            ws = wb.active
            
            # Redenumește foaia de lucru
            ws.title = base_name
            
            # Citește datele din CSV-uri
            df_lunar = pd.read_csv(files["MM"])
            df_quarter = pd.read_csv(files["YY"])
            
            num_rows = len(df_lunar)
            num_rows_quarter = len(df_quarter)
            
            # Inserăm datele 1M (exemplu: coloanele A și B)
            for index, row in df_lunar.iterrows():
                ws.cell(row=index+2, column=1, value=row.iloc[0])  # data MM
                ws.cell(row=index+2, column=2, value=row.iloc[1])  # valoare YY
            
            # Inserăm datele 3M (exemplu: coloanele R și S)
            for index, row in df_quarter.iterrows():
                if index < num_rows:  
                    ws.cell(row=index+2, column=15, value=row.iloc[0])  # data MM
                    ws.cell(row=index+2, column=16, value=row.iloc[1])  # valoare YY
            
            # Eliminăm rândurile și formulele în plus
            remove_extra_rows(ws, num_rows)
            remove_extra_rows_quarter(ws, num_rows_quarter)
            clear_last_row_formulas_lunar(ws, num_rows + 1)
            clear_last_row_formulas_quarter(ws, num_rows_quarter + 1)
            
            # Salvează fișierul XLSX rezultat
            output_file = os.path.join(output_folder, f"{base_name}.xlsx")
            wb.save(output_file)
            
            msg_ok = f"✔ Fișier completat: {output_file}"
            print(msg_ok)
            log_processing_info(log_file, msg_ok)
        
        except Exception as e:
            # Dacă apare vreo eroare la procesare, o logăm și continuăm
            error_msg = f"Eroare la procesarea '{base_name}': {e}"
            print(error_msg)
            log_processing_info(log_file, error_msg)
    
    # Log: finalizare
    log_processing_info(log_file, "Final procesare fișiere.")

# Exemplu de utilizare
if __name__ == "__main__":
    csv_folder = "csv"
    template_path = "template/momyoy.xlsx"
    output_folder = "output"
    
    # Fișierul de log (va fi creat automat dacă nu există)
    log_file = "process_log.txt"
    
    process_csv_to_xlsx(csv_folder, template_path, output_folder, log_file)
