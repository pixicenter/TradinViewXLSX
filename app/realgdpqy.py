import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import argparse

def log_processing_info(log_file, message):
    """
    Scrie în fișierul 'log_file' un mesaj, cu data/ora curentă.
    Creează fișierul dacă nu există.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{now}] {message}\n")


def clear_last_row_formulas_lunar(ws, last_data_row):
    """
    Elimină formulele din coloanele specifice pentru ultimul rând generat 
    pentru date lunare (1M).
    """
    columns_to_clear = ["J", "K", "L", "M", "N", "O", "W", "X", "Y", "Z", "AA", "AB"]
    for col in columns_to_clear:
        ws[f"{col}{last_data_row}"].value = None

def remove_extra_rows(ws, num_rows):
    """
    Elimină rândurile în plus din fișierul Excel dacă sunt mai multe decât 
    în fișierul CSV 1M (num_rows).
    """
    max_existing_rows = ws.max_row
    if max_existing_rows > num_rows + 1:
        ws.delete_rows(num_rows + 2, max_existing_rows - num_rows - 1)



def process_csv_to_xlsx(csv_folder, template_path, output_folder, log_file="process_log.txt"):
    """
    Completează template.xlsx cu datele din fișierele CSV,
    fără a modifica structura acestuia. 
    Scrie log-uri în 'log_file'.
    """
    csv_files = [f for f in os.listdir(csv_folder) if f.endswith(".csv")]
    
    # Creăm folderul de output dacă nu există
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Log: începem procesarea
    log_processing_info(log_file, f"Start procesare fișiere'")
    
    for csv_file in csv_files:
        base_name = os.path.splitext(csv_file)[0]  # Scoatem extensia
        csv_path = os.path.join(csv_folder, csv_file)
        
        try:
            # Încarcă șablonul Excel
            wb = load_workbook(template_path)
            ws = wb.active
            
            # Redenumește foaia de lucru
            ws.title = base_name
            
            # Citește datele din CSV-uri
            df = pd.read_csv(csv_path)
            num_rows = len(df)
            
            # Inserăm datele în coloanele A și B ale fișierului Excel
            for index, row in df.iterrows():
                ws.cell(row=index+2, column=1, value=row.iloc[0])  # Data CSV
                ws.cell(row=index+2, column=2, value=row.iloc[1])  # Valoare CSV
            
            # Eliminăm rândurile și formulele în plus
            remove_extra_rows(ws, num_rows)
            clear_last_row_formulas_lunar(ws, num_rows + 1)
            
            # Salvează fișierul XLSX rezultat
            output_file = os.path.join(output_folder, f"{base_name}.xlsx")
            wb.save(output_file)
            
            msg_ok = f"✔ Fișier completat: {base_name}.xlsx"
            print(msg_ok)
            log_processing_info(log_file, msg_ok)
        
        except Exception as e:
            # Dacă apare vreo eroare la procesare, o logăm și continuăm
            error_msg = f"Eroare la procesarea '{base_name}': {e}"
            print(error_msg)
            log_processing_info(log_file, error_msg)
    
    # Log: finalizare
    log_processing_info(log_file, "Final procesare fișiere.")

# ------------------------------------------------
# EXEMPLU de utilizare (script standalone):
# ------------------------------------------------
if __name__ == "__main__":

    # 1) Încarcă argumentele din linia de comandă
    parser = argparse.ArgumentParser(description="Procesare CSV în XLSX.")
    parser.add_argument(
        "--csv-folder",
        type=str,
        default=None,
        help="Folderul unde sunt fișierele CSV."
    )
    parser.add_argument(
        "--template-path",
        type=str,
        default=None,
        help="Calea către fișierul template XLSX."
    )
    parser.add_argument(
        "--output-folder",
        type=str,
        default=None,
        help="Folderul unde se vor salva fișierele XLSX rezultate."
    )
    parser.add_argument(
        "--log-file",
        type=str,
        default=None,
        help="Numele/locația fișierului de log."
    )
    args = parser.parse_args()

    # 2) Preia și din variabile de mediu (dacă nu vin din CLI):
    #    CSV_FOLDER, TEMPLATE_PATH, OUTPUT_FOLDER, LOG_FILE
    csv_folder_env = os.getenv("CSV_FOLDER", "csv")
    template_path_env = os.getenv("TEMPLATE_PATH", "template/template.xlsx")
    output_folder_env = os.getenv("OUTPUT_FOLDER", "output")
    log_file_env = os.getenv("LOG_FILE", "process_log.txt")

    # 3) Alege ordinea priorităților:
    #    CLI > Environment Variables > Valorile default
    csv_folder = args.csv_folder if args.csv_folder else csv_folder_env
    template_path = args.template_path if args.template_path else template_path_env
    output_folder = args.output_folder if args.output_folder else output_folder_env
    log_file = args.log_file if args.log_file else log_file_env

    # --------------------------------------------------------
    # Liniile următoare erau în codul original - le comentăm,
    # dar nu le ștergem (cerința ta).
    # --------------------------------------------------------
    # csv_folder = "csv"
    # template_path = "template/template.xlsx"
    # output_folder = "output"
    # log_file = "process_log.txt"
    #
    # process_csv_to_xlsx(csv_folder, template_path, output_folder, log_file)

    # --------------------------------------------------------
    # Apel final cu parametrii obținuți din CLI / ENV
    # --------------------------------------------------------
    process_csv_to_xlsx(
        csv_folder=csv_folder,
        template_path=template_path,
        output_folder=output_folder,
        log_file=log_file
    )